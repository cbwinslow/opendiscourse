from __future__ import annotations

import csv
import json
import re
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook
from psycopg.types.json import Jsonb

from ..capacity import remote_size, storage_preview
from ..config import settings
from ..contracts import get_contract
from .base import IngestionRun, client, json_response
from .bulk import ArtifactSpec, data_root, download


def ingest_acs(
    year: int, state: str, variables: list[str], dataset: str = "acs/acs5"
) -> int:
    if not settings.census_api_key:
        raise ValueError(
            "CENSUS_API_KEY is required for Census ingestion; add it to .env"
        )
    get = ["NAME", *[v for v in variables if v != "NAME"]]
    params = {"get": ",".join(get), "for": "county:*", "in": f"state:{state}"}
    params["key"] = settings.census_api_key
    url = f"https://api.census.gov/data/{year}/{dataset}"
    with (
        client() as http,
        IngestionRun(
            "census.acs_5", {"year": year, "state": state, "variables": get}
        ) as run,
    ):
        response = http.get(url, params=params)
        payload = json_response(response)
        payload_id = run.store_payload(response, payload)
        headers, *rows = payload
        for row in rows:
            record = dict(zip(headers, row, strict=True))
            geoid = f"{record['state']}{record['county']}"
            with run.conn.cursor() as cur:
                cur.execute(
                    """INSERT INTO core.geography (geography_type, geoid, name, state_fips, county_fips)
                       VALUES ('county', %s, %s, %s, %s)
                       ON CONFLICT (geography_type, geoid) DO UPDATE SET name = EXCLUDED.name
                       RETURNING geography_id""",
                    (geoid, record.get("NAME"), record["state"], record["county"]),
                )
                geography_id = cur.fetchone()["geography_id"]
                for field_id in get:
                    if field_id == "NAME" or record.get(field_id) is None:
                        continue
                    try:
                        value = float(record[field_id])
                    except ValueError:
                        value = None
                    cur.execute(
                        """INSERT INTO fact.measurement (dataset_id, field_id, geography_id, period_start, vintage_date, value_numeric, value_text, source_payload_id)
                           VALUES ('census.acs_5', %s, %s, make_date(%s, 1, 1), make_date(%s, 1, 1), %s, %s, %s)
                           ON CONFLICT (dataset_id, field_id, geography_id, period_start, period_end, vintage_date)
                           DO UPDATE SET value_numeric = EXCLUDED.value_numeric, value_text = EXCLUDED.value_text, source_payload_id = EXCLUDED.source_payload_id""",
                        (
                            field_id,
                            geography_id,
                            year,
                            year,
                            value,
                            record[field_id],
                            payload_id,
                        ),
                    )
                    run.record_count += 1
            run.conn.commit()
        return run.record_count


STATE_FIPS = (
    "01",
    "02",
    "04",
    "05",
    "06",
    "08",
    "09",
    "10",
    "11",
    "12",
    "13",
    "15",
    "16",
    "17",
    "18",
    "19",
    "20",
    "21",
    "22",
    "23",
    "24",
    "25",
    "26",
    "27",
    "28",
    "29",
    "30",
    "31",
    "32",
    "33",
    "34",
    "35",
    "36",
    "37",
    "38",
    "39",
    "40",
    "41",
    "42",
    "44",
    "45",
    "46",
    "47",
    "48",
    "49",
    "50",
    "51",
    "53",
    "54",
    "55",
    "56",
    "72",
)


def _acs_table_list_url(year: int) -> str:
    return f"https://www2.census.gov/programs-surveys/acs/tech_docs/table_shells/table_lists/{year}_DataProductList.xlsx"


def _column(headers: list[str], *names: str) -> int | None:
    normalized = [header.casefold().replace(" ", "") for header in headers]
    for name in names:
        needle = name.casefold().replace(" ", "")
        if needle in normalized:
            return normalized.index(needle)
    return None


def discover_acs_tables(year: int) -> dict[str, Any]:
    """Download only the official ACS table-list workbook and create a manifest.

    The manifest is intentionally a selection aid, not permission to download
    data. Housing candidates are transparent heuristics: B25/C25/S25 families,
    DP04, and titles containing "housing". An operator reviews the generated
    IDs in a contract before any summary-file acquisition is enabled.
    """
    url = _acs_table_list_url(year)
    preview = storage_preview([remote_size(url)])
    if not preview["approved"]:
        raise ValueError(f"Table-list metadata preflight failed: {preview['reason']}")
    spec = ArtifactSpec(
        dataset_id="census.acs_5",
        artifact_key=f"tables-{year}",
        url=url,
        filename=f"{year}_DataProductList.xlsx",
        period_start=date(year, 1, 1),
        period_end=date(year, 12, 31),
        metadata={
            "kind": "acs_table_list",
            "year": year,
            "admission": "official_metadata",
        },
    )
    with IngestionRun(
        "census.acs_5",
        {"action": "discover", "year": year, "artifact": spec.artifact_key},
        mode="plan",
    ) as run:
        path = download(spec)
        book = load_workbook(
            filename=BytesIO(path.read_bytes()), read_only=True, data_only=True
        )
        sheet = book.active
        rows = sheet.iter_rows(values_only=True)
        headers = [
            str(value).strip() if value is not None else "" for value in next(rows)
        ]
        table_col = _column(headers, "table id", "tableid")
        title_col = _column(headers, "table title", "tabletitle")
        universe_col = _column(headers, "table universe", "tableuniverse")
        product_col = _column(headers, "data product type", "dataproducttype")
        # The "Year" column is the actual 1-year/5-year *availability* signal
        # ("1", "5", or "1,5") -- confirmed live (2026-08-07) that every one
        # of this session's "unresolvable size" ACS Detailed Tables (32 of
        # 33) is simply a table this column marks 1-year-only, which the
        # table-based 5-year Summary File never publishes at all (a clean
        # 404, not a server stall). This is a *different* signal from the
        # geography-restriction columns below, which describe additional
        # exclusions *within* whichever years a table is already published.
        year_col = _column(headers, "year")
        one_year_col = _column(
            headers,
            "1-year geography restrictions\n(with summary levels in parentheses)",
        )
        five_year_col = _column(
            headers,
            "5-year geography restrictions\n(with summary levels in parentheses)",
        )
        if table_col is None or title_col is None or year_col is None:
            raise ValueError(f"Unexpected ACS table-list columns: {headers}")
        tables: list[dict[str, str]] = []
        for row in rows:
            table_id = row[table_col] if table_col < len(row) else None
            if not isinstance(table_id, str) or not table_id.strip():
                continue
            title = row[title_col] if title_col < len(row) and row[title_col] else ""
            universe = (
                row[universe_col]
                if universe_col is not None
                and universe_col < len(row)
                and row[universe_col]
                else ""
            )
            product = (
                row[product_col]
                if product_col is not None
                and product_col < len(row)
                and row[product_col]
                else ""
            )
            table_year = (
                str(row[year_col]).strip()
                if year_col < len(row) and row[year_col]
                else ""
            )
            one_year = (
                row[one_year_col]
                if one_year_col is not None
                and one_year_col < len(row)
                and row[one_year_col]
                else ""
            )
            five_year = (
                row[five_year_col]
                if five_year_col is not None
                and five_year_col < len(row)
                and row[five_year_col]
                else ""
            )
            tables.append(
                {
                    "id": table_id.strip(),
                    "title": str(title).strip(),
                    "universe": str(universe).strip(),
                    "product": str(product).strip(),
                    "year": table_year,
                    "one_year_geo_restriction": str(one_year).strip(),
                    "five_year_geo_restriction": str(five_year).strip(),
                }
            )
        book.close()
        housing = [
            table
            for table in tables
            if re.match(r"^(B25|C25|S25)", table["id"], flags=re.IGNORECASE)
            or table["id"].upper() == "DP04"
            or "housing" in table["title"].casefold()
        ]
        manifest = {
            "version": 1,
            "provider": "census",
            "release_year": year,
            "source_url": url,
            "artifact": str(path),
            "tables": tables,
            "housing_candidates": housing,
            "selection_rule": "B25/C25/S25 prefixes, DP04, or title containing housing; review before approval",
        }
        output = data_root().parent / "meta" / "acs" / str(year) / "tables.json"
        output.parent.mkdir(parents=True, exist_ok=True)
        temp = output.with_suffix(".json.part")
        temp.write_text(json.dumps(manifest, indent=2, sort_keys=True) + "\n")
        temp.replace(output)
        run.record_count = len(tables)
    return {
        "year": year,
        "table_count": len(tables),
        "housing_candidate_count": len(housing),
        "manifest": str(output),
        "next": "Review housing_candidates, copy approved IDs into a contract, then create a bulk file and storage plan.",
    }


def load_acs_field_catalog(year: int) -> int:
    """Download this release's official Table Shells and populate catalog.dataset_field.

    catalog.dataset_field exists in the schema but has no other writer
    anywhere in this codebase; without it, a loaded field_id like
    B19013_E001 is meaningless without cross-referencing Census
    documentation by hand. The shells file lists every table's every
    variable with a human-readable label/title/universe; one shell line
    becomes two dataset_field rows (the _E estimate and _M margin-of-error
    variants actually used in fact.acs_bulk_estimate.field_id).
    """
    base = f"https://www2.census.gov/programs-surveys/acs/summary_file/{year}/table-based-SF"
    url = f"{base}/documentation/ACS{year}5YR_Table_Shells.txt"
    preview = storage_preview([remote_size(url)])
    if not preview["approved"]:
        raise ValueError(f"Table shells preflight failed: {preview['reason']}")
    spec = ArtifactSpec(
        dataset_id="census.acs_5_bulk",
        artifact_key=f"acs5-{year}-table-shells",
        url=url,
        filename=f"ACS{year}5YR_Table_Shells.txt",
        period_start=date(year, 1, 1),
        period_end=date(year, 12, 31),
        metadata={"kind": "table_shells", "year": year},
    )
    with IngestionRun(
        "census.acs_5_bulk",
        {"action": "field_catalog", "year": year, "artifact": spec.artifact_key},
        mode="plan",
    ) as run:
        path = download(spec)
        valid_from = date(year, 1, 1)
        rows: list[tuple[Any, ...]] = []
        with path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle, delimiter="|")
            for record in reader:
                unique_id = (record.get("Unique ID") or "").strip()
                table_id, sep, number = unique_id.partition("_")
                if not sep or not table_id or not number:
                    continue
                label = (record.get("Label") or "").strip()
                title = (record.get("Title") or "").strip()
                universe = (record.get("Universe") or "").strip()
                data_type = (record.get("Type") or "").strip()
                description = f"{title} -- {universe}" if universe else title
                metadata = {
                    "table_id": table_id,
                    "title": title,
                    "universe": universe,
                    "line": (record.get("Line") or "").strip(),
                    "indent": (record.get("Indent") or "").strip(),
                }
                for measure in ("E", "M"):
                    rows.append(
                        (
                            "census.acs_5_bulk",
                            f"{table_id}_{measure}{number}",
                            label,
                            data_type,
                            description,
                            valid_from,
                            Jsonb(metadata),
                        )
                    )
        with run.conn.cursor() as cur:
            for start in range(0, len(rows), 2_000):
                cur.executemany(
                    """INSERT INTO catalog.dataset_field
                       (dataset_id, field_id, label, data_type, description, valid_from, metadata)
                       VALUES (%s, %s, %s, %s, %s, %s, %s)
                       ON CONFLICT (dataset_id, field_id, valid_from) DO UPDATE SET
                         label = EXCLUDED.label, data_type = EXCLUDED.data_type,
                         description = EXCLUDED.description, metadata = EXCLUDED.metadata""",
                    rows[start : start + 2_000],
                )
        run.conn.commit()
        run.record_count = len(rows)
    return run.record_count


def search_acs_tables(
    year: int, text: str, product: str | None = None, limit: int = 50
) -> dict[str, Any]:
    """Search the local official ACS catalog using IDs, titles, and universes."""
    manifest_path = data_root().parent / "meta" / "acs" / str(year) / "tables.json"
    if not manifest_path.is_file():
        raise FileNotFoundError(
            f"No ACS manifest for {year}; run census-discover first"
        )
    terms = [term.casefold() for term in text.split() if term]
    product_filter = product.casefold() if product else None
    tables = json.loads(manifest_path.read_text())["tables"]
    matches = [
        table
        for table in tables
        if all(
            term
            in " ".join(
                str(table.get(key, ""))
                for key in ("id", "title", "universe", "product")
            ).casefold()
            for term in terms
        )
        and (
            product_filter is None
            or product_filter in table.get("product", "").casefold()
        )
    ]
    matches.sort(key=lambda table: table["id"])
    return {
        "year": year,
        "query": text,
        "match_count": len(matches),
        "results": matches[:limit],
        "truncated": len(matches) > limit,
    }


# Quality-measure/allocation-flag tables carry no substantive data.
_ACS_EXCLUDE_PREFIX = ("B98", "B99")
# Race/ethnicity-iterated variants of a base table (e.g. B01001A..I = "Sex
# by Age" x9 race/ethnicity groups) were excluded through 2026-08-07 on the
# claim that they add "zero new information beyond B02/B03" -- that claim
# was wrong and has been retracted: iterations are genuine cross-tabulations
# (e.g. B19013A-I is median household income BY race, B25003A-I is tenure
# BY race) not recoverable from the marginal B02/B03 tables, matching how
# NHGIS itself treats these as first-class "breakdowns" rather than
# redundant. No longer excluded by default -- see relevant_acs_tables().

# Confirmed live (2026-08-07): every table this session found "unresolvable"
# via a size probe -- except B08009 -- is simply a table the official
# DataProductList.xlsx (downloaded per-year by discover_acs_tables(), see
# the "year" manifest field) marks 1-year-only. The table-based 5-year
# Summary File never publishes 1-year-only tables at all (a clean 404, not
# a server stall or a probing bug), and B25142/B25143 in particular don't
# even exist in earlier years' manifests -- Census introduced those table
# IDs starting with the 2024 release. relevant_acs_tables() now excludes
# 1-year-only tables directly via that "year" field, so this dict no
# longer needs the ~30 previously-hand-maintained IDs per year.
#
# B08009 (2024 only) is the one genuine exception: the xlsx marks it
# "1,5" (nominally published at 5-year), yet a direct curl returns a
# real Census-branded 404 page, not a WAF artifact and not a timeout --
# Census's own catalog appears to be wrong about this one table. Keep
# this dict for exactly this kind of case: a table relevant_acs_tables()
# would otherwise select but that a size probe still can't resolve.
#
# Separately (not covered by this dict, since it isn't a per-year/per-table
# constant): a scan of every downloaded 2021-2024 file found 13 that are
# an HTML "Request Rejected" WAF page instead of real data, staged to zero
# rows so no bad data reached fact.acs_bulk_estimate -- but 12 of the 13
# ARE genuinely 5-year-available tables (only B18121 is explained by the
# 1-year-only issue above) and are missing from the loaded dataset purely
# because of a download-layer reliability gap (likely rate-limiting from
# this session's request volume). Those need a straightforward re-download,
# not a filter change -- see docs/census-bulk-roadmap.md.
ACS_SIZE_PROBE_UNRESOLVABLE: dict[int, frozenset[str]] = {
    2024: frozenset({"B08009"}),
}


def relevant_acs_tables(year: int, include_housing_detail: bool = True) -> list[str]:
    """Return a reviewable, regenerable relevance-filtered ACS Detailed Table list.

    Excludes only what carries no additional information beyond what's
    already selected: quality/allocation-flag tables (B98/B99, counts of
    imputed cells rather than estimates) and C-prefix "collapsed" tables
    (a strictly less-detailed duplicate of the corresponding B-table,
    which is already included). Everything else -- including the full B25
    housing family, the narrow administrative families (B10/B13/B26/B29)
    that were excluded in an earlier, narrower pass, and race/ethnicity
    iteration tables (e.g. B01001A..I, B19013A..I) that were also
    excluded in an earlier pass on a since-retracted "adds zero new
    information" claim -- is included by default per an explicit
    "comprehensive coverage" decision (2026-08-07): pass
    include_housing_detail=False to fall back to excluding B25 if a
    narrower housing scope is ever wanted again.

    Also excludes tables the manifest's ``year`` field marks 1-year-only
    (no "5" in a comma-separated "1"/"5"/"1,5" value) -- confirmed live
    (2026-08-07) that this is the true explanation for every "unresolvable
    size" table found this session except one (B08009): the table-based
    5-year Summary File never publishes 1-year-only tables at all, so
    they 404 rather than the size probe finding anything. Manifests
    generated before this field existed have an empty ``year`` string for
    every table, which this treats as "unknown, don't exclude" to avoid
    silently dropping tables from a stale manifest -- rerun
    ``census-discover`` to get real per-table year availability.
    """
    manifest_path = data_root().parent / "meta" / "acs" / str(year) / "tables.json"
    if not manifest_path.is_file():
        raise FileNotFoundError(
            f"No ACS manifest for {year}; run census-discover first"
        )
    tables = json.loads(manifest_path.read_text())["tables"]
    selected = []
    for table in tables:
        table_id = table["id"]
        if table.get("product") != "Detailed Table":
            continue
        if table_id.startswith("C"):
            continue
        if table_id[:3] in _ACS_EXCLUDE_PREFIX:
            continue
        if not include_housing_detail and table_id[:3] == "B25":
            continue
        year_field = table.get("year", "")
        if year_field and "5" not in year_field.split(","):
            continue
        selected.append(table_id)
    return sorted(selected)


def _acs_group_endpoint(product: str) -> str:
    if product == "Detailed Table":
        return "acs/acs5"
    if product == "Data Profile":
        return "acs/acs5/profile"
    if product.startswith("Subject Tables"):
        return "acs/acs5/subject"
    raise ValueError(f"No Census group-metadata endpoint is configured for {product!r}")


def describe_acs_table(
    year: int, table_id: str, include_annotations: bool = False
) -> dict[str, Any]:
    """Retrieve and record field-level metadata for one locally cataloged table."""
    catalog = search_acs_tables(year, table_id, limit=10_000)["results"]
    match = next(
        (table for table in catalog if table["id"].casefold() == table_id.casefold()),
        None,
    )
    if match is None:
        raise ValueError(f"{table_id!r} is not in the {year} ACS table-list manifest")
    endpoint = _acs_group_endpoint(match["product"])
    with (
        client() as http,
        IngestionRun(
            "census.acs_5",
            {
                "action": "describe",
                "year": year,
                "table": match["id"],
                "endpoint": endpoint,
            },
            mode="plan",
        ) as run,
    ):
        response = http.get(
            f"https://api.census.gov/data/{year}/{endpoint}/groups/{match['id']}.json"
        )
        payload = json_response(response)
        payload_id = run.store_payload(response, payload)
        available = [
            {
                "id": field_id,
                "label": metadata.get("label"),
                "concept": metadata.get("concept"),
                "type": metadata.get("predicateType"),
            }
            for field_id, metadata in payload.get("variables", {}).items()
            if field_id not in {"NAME", "GEO_ID"}
        ]
        fields = (
            available
            if include_annotations
            else [
                field for field in available if not field["id"].endswith(("EA", "MA"))
            ]
        )
        fields.sort(key=lambda field: field["id"])
        run.record_count = len(fields)
    return {
        "table": match,
        "field_count": len(fields),
        "available_field_count": len(available),
        "fields": fields,
        "payload_id": payload_id,
    }


def review_bulk_contract(contract_id: str) -> dict[str, Any]:
    """Resolve a draft ACS bulk selection against its discovered table manifest."""
    contract = get_contract(contract_id)
    if contract.get("kind") != "acs_bulk":
        raise ValueError(f"{contract_id!r} is not an ACS bulk contract")
    year = contract["year"]
    manifest_path = data_root().parent / "meta" / "acs" / str(year) / "tables.json"
    if not manifest_path.is_file():
        raise FileNotFoundError(
            f"No ACS manifest for {year}; run census-discover first"
        )
    manifest = json.loads(manifest_path.read_text())
    selection = contract["selection"]
    prefixes = tuple(item.upper() for item in selection.get("include_prefixes", []))
    include_ids = {item.upper() for item in selection.get("include_ids", [])}
    add_ids = {item.upper() for item in selection.get("add_ids", [])}
    exclude_ids = {item.upper() for item in selection.get("exclude_ids", [])}
    terms = [item.casefold() for item in selection.get("include_title_terms", [])]
    product_types = {
        "detailed_5": "Detailed Table",
        "profile_5": "Data Profile",
        "subject_5": "Subject Tables",
    }
    accepted_products = tuple(product_types[item] for item in contract["products"])
    selected = []
    for table in manifest["tables"]:
        table_id = table["id"].upper()
        title = table["title"].casefold()
        included = (
            table_id.startswith(prefixes)
            or table_id in include_ids
            or table_id in add_ids
            or any(term in title for term in terms)
        )
        included = included and any(
            table.get("product", "").startswith(kind) for kind in accepted_products
        )
        if included and table_id not in exclude_ids:
            selected.append(table)
    selected.sort(key=lambda table: table["id"])
    return {
        "contract": contract_id,
        "enabled": contract.get("enabled", False),
        "approval": contract.get("approval"),
        "year": year,
        "manifest": str(manifest_path),
        "selected_count": len(selected),
        "selected_ids": [table["id"] for table in selected],
        "excluded_ids": sorted(exclude_ids),
        "next": "Edit include/exclude/add IDs in the contract. It remains disabled until an exact file manifest and capacity approval exist.",
    }


def plan_contract(contract_id: str) -> dict[str, Any]:
    """Discover selected Census table metadata and persist a no-data load plan.

    This deliberately reads only group definitions. It neither calls a data
    endpoint nor writes measurements, so operators can review the exact fields
    and request size before enabling a load.
    """
    contract = get_contract(contract_id)
    if contract["provider"] != "census":
        raise ValueError(f"{contract_id!r} is not a Census contract")

    endpoint = contract["endpoint"].strip("/")
    suffix = contract.get("estimate_suffix", "E")
    groups: list[dict[str, Any]] = []
    with (
        client() as http,
        IngestionRun(
            contract["dataset"],
            {"contract": contract_id, "action": "plan", "year": contract["year"]},
            mode="plan",
        ) as run,
    ):
        for group_id in contract["groups"]:
            response = http.get(
                f"https://api.census.gov/data/{contract['year']}/{endpoint}/groups/{group_id}.json"
            )
            payload = json_response(response)
            payload_id = run.store_payload(response, payload)
            variables = payload.get("variables", {})
            selected = sorted(
                key for key in variables if key != "NAME" and key.endswith(suffix)
            )
            groups.append(
                {
                    "id": group_id,
                    "title": payload.get("description") or payload.get("title"),
                    "variables": len(selected),
                    "sample": selected[:5],
                    "payload_id": payload_id,
                }
            )
        variable_count = sum(group["variables"] for group in groups)
        requests = len(contract["states"]) * ((variable_count + 44) // 45)
        run.record_count = len(groups)
    return {
        "contract": contract_id,
        "dataset": contract["dataset"],
        "year": contract["year"],
        "geography": contract["geography"],
        "states": contract["states"],
        "groups": groups,
        "variable_count": variable_count,
        "estimated_requests": requests,
        "next": "Review this plan; use the existing explicit ACS bootstrap only after approval.",
    }


def bootstrap_housing(year: int, states: list[str]) -> int:
    """Load curated ACS housing table groups in API-safe variable batches."""
    if not settings.census_api_key:
        raise ValueError(
            "CENSUS_API_KEY is required for Census ingestion; add it to .env"
        )
    manifest = yaml.safe_load(
        (
            Path(__file__).resolve().parents[3]
            / "inventory"
            / "acs_housing_groups.yaml"
        ).read_text()
    )
    total = 0
    with client() as http:
        for group in manifest["groups"]:
            response = http.get(
                f"https://api.census.gov/data/{year}/acs/acs5/groups/{group}.json",
                params={"key": settings.census_api_key},
            )
            metadata = json_response(response)
            variables = sorted(
                variable_id
                for variable_id in metadata["variables"]
                if variable_id.endswith("E")
            )
            for start in range(0, len(variables), 45):
                for state in states:
                    total += ingest_acs(year, state, variables[start : start + 45])
    return total
